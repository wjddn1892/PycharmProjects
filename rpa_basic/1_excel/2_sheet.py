from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()  # 새로운 sheet 기본 이름으로 생성
ws.title = "MySheet"    # sheet 이름 변경
ws.sheet_properties.tabColor = "99ffcc"   # RGB 형태로 값을 넣어주면 탭 색상 변경

# sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)    # 2번째 index 에 Sheet 생성

new_ws = wb["NewSheet"] # 딕셔너리 형태로 sheet 에 접근

print(wb.sheetnames)    # 모든 sheet 이름 확인

# sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")