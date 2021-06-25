from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):  # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 column 만 가져오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"]   # 영어, 수학 column 함께 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]   # 1번째 row 만 가져오기
for cell in row_title:
    print(cell.value)

# row_range = ws[2:6] # 2번째 줄(row)에서 6번째 줄까지 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row]    # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ") # A10, AZ250
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")  # end=" " 는 줄바꿈 역할
        print(xy[0], end="")    # A
        print(xy[1], end=" ")   # 1
    print()

# # 전체 rows
# # print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)
#
# # 전체 columns
# # print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)
#
# for row in ws.iter_rows():  # 전체 row
#     print(row[1].value)
#
# for column in ws.iter_cols():  # 전체 col
#     print(column[0].value)

for row in ws.iter_rows(min_row=1, max_row=5):
    print(row[1].value)

wb.save("sample.xlsx")